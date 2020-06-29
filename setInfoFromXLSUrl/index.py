#!/usr/bin/env python
import http.client
import mimetypes
import json
import openpyxl
# https://lcdns-vv.learnta.com/dyJM98O-x45NBZz6kDTp62JhcQ4=/lie3xXsEklK0bx3iaLPEeAAZGGbe
book = openpyxl.load_workbook('video.xlsx')

sheet = book.active
for index in range(729, 735):

    a3 = sheet.cell(row=index+1, column=3)

    # print(str(a3.value).split('/')[2])
    conn = http.client.HTTPSConnection(str(a3.value).split('/')[2])
    payload = ''
    headers = {}

    # print(str(a3.value).replace('https://' +
    #                             str(a3.value).split('/')[2], '')+"?avinfo")
    conn.request(
        "GET", str(a3.value).replace('https://' +
                                     str(a3.value).split('/')[2], '')+"?avinfo", payload, headers)
    res = conn.getresponse()
    data = res.read()
    # print(data.decode("utf-8"))
    y = json.loads(data.decode("utf-8"))

    # the result is a Python dictionary:
    print(y["format"]["duration"])

    # print(a3.value)
    sheet.cell(row=index+1, column=4).value = y["format"]["duration"]

    pass


book.save("sample.xlsx")
