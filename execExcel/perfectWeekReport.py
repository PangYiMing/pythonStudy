#!/usr/bin/env python3.7
# coding:utf-8
import openpyxl
import os
import sys
import time
import shutil

# 1 先把目标文件复制到目标文件夹下面
# 2 处理需求1 分列
# 3 处理需求2 过滤 删除
# 4 处理需求2 覆盖


def getTargetPathByName(name):
    targetFile = '周报/' + name
    path = sys.argv[0].replace('perfectWeekReport.py', targetFile)
    return path


def getBuildPath():
    bPath = getTargetPathByName(
        '') + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    # 格式化成2016-03-20 11:45:39形式
    return bPath

# 如果存在直接返回true，不存在就创建，返回创建结果


def mkdir(path):
    isExists = os.path.exists(path)

    if not isExists:
        os.makedirs(path)
        isExists = os.path.exists(path)
        resolve = '成功' if isExists else '失败'
        print(path+'创建'+resolve)
        return isExists
    return True

# 获取整行


def getTargetRow(ws, id, amount=None):
    nextId = id+1
    rows = ws.iter_rows(id, nextId)
    for row in rows:
        return row
    return None

# 获取整列


def getTargetCol(ws, id, amount=None):
    nextId = id+1
    cols = ws.iter_cols(id, nextId)
    for col in cols:
        return col
    return None


def getColByTitle(ws, title):
    col = None
    row = getTargetRow(ws, 0)
    for cell in row:
        if cell.value == title:
            col = getTargetCol(ws, cell.column)
            break
        pass
    return col


def reverseCombiner(rowList):
    # Don't do anything for empty list. Otherwise,
    # make a copy and sort.

    if len(rowList) == 0:
        return []
    sortedList = rowList[:]
    sortedList.sort()

    # Init, empty tuple, use first item for previous and
    # first in this run.

    tupleList = []
    firstItem = sortedList[0]
    prevItem = sortedList[0]

    # Process all other items in order.

    for item in sortedList[1:]:
        # If start of new run, add tuple and use new first-in-run.

        if item != prevItem + 1:
            tupleList = [(firstItem, prevItem + 1 - firstItem)] + tupleList
            firstItem = item

        # Regardless, current becomes previous for next loop.

        prevItem = item

    # Finish off the final run and return tuple list.

    tupleList = [(firstItem, prevItem + 1 - firstItem)] + tupleList
    return tupleList

# Test data, hit me with anything :-)

# myList = [1, 70, 71, 72, 98, 21, 22, 23, 24, 25, 99]

# Create tuple list, show original and that list, then process.

# tuples = reverseCombiner(myList)
# print(f"Original: {myList}")
# print(f"Tuples:   {tuples}\n")
# for tuple in tuples:
#     print(f"Would execute: worksheet.delete_rows({tuple[0]}, {tuple[1]})")


def execData(path, titlePath, path2):
    print('execData:start load'+',time :' +
          time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    book = openpyxl.load_workbook(path)

    ws = book.active
    print('execData:start exec'+',time :' +
          time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    # 拿到工单分类名称列 根据名字获取整列
    col = getColByTitle(ws, '工单分类名称')
    # 业务类型的column的id
    ywlxId = col[0].column+5
    # 工单分类名称列后加4列 4列问题 1列业务类型 增加2列 省/战区
    ws.insert_cols(col[0].column+1, 7)
    print('execData: add col'+',time :' +
          time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    # 新增col以后，更新row
    row = getTargetRow(ws, 0)

    num = ['零', '一', '二', '三', '四', '五', '六', '七', '八', '九']
    secondProbleColIndex = row[col[0].column+1].column
    provinceIndex = row[col[0].column+5].column

    # 设置 这四列到名称
    for index in range(0, 7):
        if index == 4:
            row[col[0].column+index].value = '业务类型'
            pass
        if index == 5:
            row[col[0].column+index].value = '省份'
            pass
        if index == 6:
            row[col[0].column+index].value = '战区'
            pass
        if index < 4:
            row[col[0].column+index].value = num[index+1]+'级问题'
            pass
        pass
    print('execData: set col name'+',time :' +
          time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    for cell in col:
        # print('execData: cell'+cell.value+',time :' +
        #       time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        if str(cell.value) != 'None':
            sArr = cell.value.split('/')
            index = 0
            for string in sArr:
                if(index >= 4):
                    break
                if cell.value != '工单分类名称':
                    ws.cell(cell.row, cell.column+index+1).value = string
                    pass
                # 业务类型 字段为次日达：二级问题为 次日达
                if index == 1:
                    if string == '次日达':
                        ws.cell(cell.row, cell.column+5).value = '次日达'
                        pass
                    else:
                        ws.cell(cell.row, cell.column+5).value = '及时达'
                        pass
                    pass
                if index == 2 and len(sArr) == 3:
                    ws.cell(cell.row, cell.column+4).value = string
                    pass
                index += 1

                pass
            pass
        pass
    print('execData: update bussiness type and problem field'+',time :' +
          time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    # 获取为 创建人部门 的列
    col = getColByTitle(ws, '创建人部门')
    print('execData: get 创建人部门 col'+',time :' +
          time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))

    # 删除操作过于消耗性能，采用复制
    titleBook = openpyxl.load_workbook(titlePath)
    titleWs = titleBook.active
    # 当 创建人部门 字段为 营销部(15000089)/客服中心二线组(15035424)
    # 删除 二级问题 为 无效咨询 的数据
    # 删除 业态 为 永辉生活 全球潮物 的数据
    ytcol = getColByTitle(ws, '业态')
    mySet = set()
    for cell in col[::-1]:
        if cell.value == '营销部(15000089)/客服中心二线组(15035424)':
            # 更新
            secondProbleStr = ws.cell(cell.row, secondProbleColIndex).value
            if secondProbleStr == '无效咨询':
                # 删除这一行
                mySet.add(cell.row)
                # ws.delete_rows(cell.row)
                # print('execData: del invalided data'+secondProbleStr+',time :' +
                #       time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
            elif '永辉生活' in str(ytcol[cell.row-1].value) or '全球潮物' in str(ytcol[cell.row-1].value):
                mySet.add(cell.row)
                # ws.delete_rows(cell.row)
                # print('execData: del invalided data'+str(ytcol[cell.row-1].value)+',time :' +
                #       time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
            pass
        pass
    for row in ws.iter_rows():
        if row[0].row != 1 and not row[0].row in mySet:
            # if not row[0].row in mySet:   # filter on first column with value 16
            titleWs.append((cell.value for cell in row))
            pass

    # print((cell.value for cell in row).values())
    # tuples = reverseCombiner(myList)
    # for tuple in tuples:
    #     print(f"Would execute: worksheet.delete_rows({tuple[0]}, {tuple[1]})")
    print('execData: del invalided data'+',time :' +
          time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    # 业务类型 字段为 全球购：业务中含有 全球潮购
    ytcol = getColByTitle(titleWs, '业态')
    for ytcell in ytcol[::-1]:
        if '全球潮物' in str(ytcell.value):
            titleWs.cell(ytcell.row, ywlxId).value = '全球购'
            pass
        pass
    print('execData: update bussiness type field'+',time :' +
          time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    # TODO 获取另外的book 城市
    book2 = openpyxl.load_workbook(path2)

    ws2 = book2.active
    cityCol = getColByTitle(titleWs, '城市')
    cityCol2 = getColByTitle(ws2, '城市-简')
    province = getColByTitle(ws2, '省份')
    area = getColByTitle(ws2, '战区')
    for cell in cityCol:
        for idx, cell2 in enumerate(cityCol2):
            if(cell2.value in str(cell.value)):
                titleWs.cell(
                    cell.row, provinceIndex).value = province[idx].value
                titleWs.cell(cell.row, provinceIndex+1).value = area[idx].value
                pass
            pass
        pass
    book2.close()
    print('execData: update province field'+',time :' +
          time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    # 保存文件
    bPath = getBuildPath()
    bfile = bPath + '/' + path.split('/')[-1]
    # 创建文件夹
    mkdir(bPath)
    # 创建文件
    book.close()
    titleBook.save(bfile)
    print('execData: save file success, path:'+bfile+',time :' +
          time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    pass


def execDataGetTitle(name):
    nameSplit = name.split('.')
    excelTitleFilePath = getTargetPathByName(
        nameSplit[0]+'Title.'+nameSplit[1])

    if not os.path.exists(excelTitleFilePath):
        print('execData:start load'+',time :' +
              time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        book = openpyxl.load_workbook(getTargetPathByName(name))

        ws = book.active
        print('execData:start exec'+',time :' +
              time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        # 拿到工单分类名称列 根据名字获取整列
        col = getColByTitle(ws, '工单分类名称')
        # 工单分类名称列后加4列 4列问题 1列业务类型 增加2列 省/战区
        ws.insert_cols(col[0].column+1, 7)
        print('execData: add col'+',time :' +
              time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        # 新增col以后，更新row
        row = getTargetRow(ws, 0)

        num = ['零', '一', '二', '三', '四', '五', '六', '七', '八', '九']
        # 设置 这四列到名称
        for index in range(0, 7):
            if index == 4:
                row[col[0].column+index].value = '业务类型'
                pass
            if index == 5:
                row[col[0].column+index].value = '省份'
                pass
            if index == 6:
                row[col[0].column+index].value = '战区'
                pass
            if index < 4:
                row[col[0].column+index].value = num[index+1]+'级问题'
                pass
            pass
        print('execData: set col name'+',time :' +
              time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        ws.delete_rows(ws.min_row+1, ws.max_row-1)
        book.close()
        book.save(excelTitleFilePath)
        pass

    return excelTitleFilePath


def copyFileToDefaultPathWithExec(name, name2):
    print('start time :'+time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    #  通过周报下面的文件名字，获取目标文件

    titlePath = execDataGetTitle(name)
    path = getTargetPathByName(name)
    path2 = getTargetPathByName(name2)
    print('titlePath:'+titlePath)
    print('path:'+path)
    print('path2:'+path2)
    # 处理数据
    execData(path, titlePath, path2)

    pass


copyFileToDefaultPathWithExec('客诉工单.xlsx', '区域信息.xlsx')
